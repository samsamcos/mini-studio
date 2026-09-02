"""
Auto-pipeline — port 9530
7-step pipeline: STT → Vision → Groq Brain → Slice → XTTS → Channel Export
+ Blueprints, Channels (50), Auto Shorts, Telegram conversation, Excel log
"""
import os, json, uuid, subprocess, threading, time, base64, re, shutil
from pathlib import Path
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename

app = Flask(__name__, static_folder=".")

# GPU encoder detection — uses NVENC on P4 if available, falls back to libx264
def _check_nvenc():
    try:
        r = subprocess.run(
            ['ffmpeg','-hide_banner','-f','lavfi','-i','nullsrc=s=16x16','-t','0.1',
             '-c:v','h264_nvenc','-f','null','-'],
            capture_output=True, timeout=8)
        return r.returncode == 0
    except Exception:
        return False
USE_NVENC = _check_nvenc()
VENC = ['-c:v','h264_nvenc','-preset','p4','-rc','vbr','-cq','23','-b:v','0'] if USE_NVENC else ['-c:v','libx264','-preset','fast','-crf','23']
print(f"[studio] GPU encoder: {'h264_nvenc (P4)' if USE_NVENC else 'libx264 (CPU)'}")

CORS(app)

WORK   = Path("/opt/studio/auto_work");  WORK.mkdir(parents=True, exist_ok=True)
EXPORT = Path("/opt/studio/media/exports"); EXPORT.mkdir(parents=True, exist_ok=True)
UPLOAD = Path("/opt/studio/media/uploads"); UPLOAD.mkdir(parents=True, exist_ok=True)

DATA_DIR        = Path("/opt/studio")
BLUEPRINTS_FILE = DATA_DIR / "blueprints.json"
CHANNELS_FILE   = DATA_DIR / "channels.json"
CHANNELS_ROOT   = DATA_DIR / "media/channels"
EXCEL_LOG       = DATA_DIR / "pipeline_log.xlsx"
DRAFTS_DIR      = DATA_DIR / "drafts"; DRAFTS_DIR.mkdir(parents=True, exist_ok=True)
PROJECTS_ROOT   = DATA_DIR / "media/projects"; PROJECTS_ROOT.mkdir(parents=True, exist_ok=True)
PROJECTS_FILE   = DATA_DIR / "projects.json"
SETTINGS_FILE   = DATA_DIR / "settings.json"
PUBLISH_QUEUE   = DATA_DIR / "publish_queue.json"
UPLOAD_QUEUE    = DATA_DIR / "upload_queue.json"

def queue_upload(local_path, remote_path):
    q = load_json(UPLOAD_QUEUE)
    q = q if isinstance(q, list) else []
    q.append({"local": str(local_path), "remote": remote_path,
              "queued": time.strftime("%Y-%m-%d %H:%M")})
    Path(UPLOAD_QUEUE).write_text(json.dumps(q, indent=2))

def upload_catchup_worker():
    """Every 5 min: if internet is back, push queued Drive uploads (oldest first)."""
    while True:
        time.sleep(300)
        try:
            q = load_json(UPLOAD_QUEUE)
            q = q if isinstance(q, list) else []
            if not q:
                continue
            probe = subprocess.run(["rclone", "lsd", "gdrive:", "--max-depth", "1"],
                                   capture_output=True, timeout=30)
            if probe.returncode != 0:
                continue    # still offline
            done, remain = [], []
            for item in q:
                if not Path(item["local"]).exists():
                    continue    # gone — drop silently
                r = subprocess.run(["rclone", "copyto", item["local"],
                                    f"gdrive:{item['remote']}", "--bwlimit", "5M"],
                                   capture_output=True, timeout=3600)
                (done if r.returncode == 0 else remain).append(item)
            Path(UPLOAD_QUEUE).write_text(json.dumps(remain, indent=2))
            if done:
                tg(f"\U0001f4f6 <b>Back online — caught up {len(done)} file(s) to "
                   f"Google Drive.</b>" +
                   ("" if not remain else f" ({len(remain)} still pending)"))
        except Exception as e:
            print(f"[catchup] {e}")

threading.Thread(target=upload_catchup_worker, daemon=True).start()
PROJECT_SUBS    = ("a-roll", "b-roll", "pics", "other")
VIDEO_EXTS      = {".mp4", ".mov", ".avi", ".mkv", ".webm", ".m4v"}
PACKAGES_DIR    = DATA_DIR / "media/packages"; PACKAGES_DIR.mkdir(parents=True, exist_ok=True)
RECALL_DIR      = DATA_DIR / "media/recall"; RECALL_DIR.mkdir(parents=True, exist_ok=True)
ASSET_INDEX     = DATA_DIR / "asset_index.json"
GDRIVE_ARCHIVE  = "mini-studio/archive"   # remote path prefix (under gdrive:)

# ── AI routing ───────────────────────────────────────────────────────────────
GROQ_KEY      = os.getenv("GROQ_API_KEY", "")
OMNIROUTE_KEY = os.getenv("OMNIROUTE_API_KEY", "")
GEMINI_KEY    = os.getenv("GEMINI_API_KEY", "")
GROQ_BASE      = "https://api.groq.com/openai/v1"
OMNIROUTE_BASE = os.getenv("OMNIROUTE_BASE_URL", "https://omniroute.online/v1")
GEMINI_BASE    = "https://generativelanguage.googleapis.com/v1beta/openai"

WHISPER_URL  = "http://localhost:8421"
TTS_URL      = "http://localhost:8422"
CHANNEL_URL  = "http://localhost:9510"

VISION_MODEL = "meta-llama/llama-4-scout-17b-16e-instruct"
BRAIN_MODEL  = "llama-3.3-70b-versatile"

# ── Telegram ─────────────────────────────────────────────────────────────────
TG_TOKEN   = os.getenv("TELEGRAM_BOT_TOKEN", "")
TG_CHAT    = os.getenv("TELEGRAM_CHAT_ID",   "")
TG_BASE    = f"https://api.telegram.org/bot{TG_TOKEN}"
_tg_offset = 0          # Telegram getUpdates offset

# ── First-run setup check ────────────────────────────────────────────────────
def _print_setup_guide():
    missing = []
    if not GROQ_KEY:  missing.append("GROQ_API_KEY")
    if not TG_TOKEN:  missing.append("TELEGRAM_BOT_TOKEN")
    if not TG_CHAT:   missing.append("TELEGRAM_CHAT_ID")
    if not missing:
        return
    border = "=" * 62
    print(f"\n+{border}+")
    print( "|          MINI STUDIO -- SETUP REQUIRED                      |")
    print(f"+{border}+\n")
    print("Add the following to /opt/studio/.env (or set as env vars):\n")
    checks = [
        ("TELEGRAM_BOT_TOKEN", "your bot token",  "@BotFather on Telegram -> /newbot"),
        ("TELEGRAM_CHAT_ID",   "your chat ID",    "@userinfobot on Telegram -> copy the id number"),
        ("GROQ_API_KEY",       "your Groq key",   "console.groq.com -> API Keys -> Create"),
        ("GEMINI_API_KEY",     "your Gemini key", "aistudio.google.com -> Get API key  (optional)"),
        ("OMNIROUTE_API_KEY",  "your key",        "omniroute.online -> Account          (optional)"),
    ]
    for key, placeholder, where in checks:
        marker = "  [X]" if key in missing else "  [OK]"
        print(f"{marker}  {key}=<{placeholder}>")
        print(f"         -> {where}")
    print()
    print("  The studio will start but Telegram and AI will not work")
    print("  until the missing keys are added and the service restarted.")
    print(f"\n  Restart:  systemctl restart mini-studio-auto")
    print(f"+{border}+\n")

_print_setup_guide()
_tg_lock   = threading.Lock()
# pending replies: job_id → threading.Event, reply text stored here when received
_pending   = {}          # jid -> {"event": Event, "reply": str or None}

def tg(msg, reply_markup=None):
    """Send a Telegram message. Non-blocking, swallows errors."""
    if not TG_TOKEN or not TG_CHAT:
        return
    try:
        import urllib.request as _ur
        payload = {"chat_id": TG_CHAT, "text": msg, "parse_mode": "HTML"}
        if reply_markup:
            payload["reply_markup"] = reply_markup
        body = json.dumps(payload).encode()
        req = _ur.Request(f"{TG_BASE}/sendMessage", data=body,
                          headers={"Content-Type": "application/json"})
        _ur.urlopen(req, timeout=10)
    except Exception as e:
        print(f"[tg] {e}")

def tg_poll():
    """Background thread: polls Telegram and resolves pending job waits."""
    global _tg_offset
    import urllib.request as _ur
    while True:
        try:
            url = f"{TG_BASE}/getUpdates?offset={_tg_offset}&timeout=30"
            with _ur.urlopen(url, timeout=40) as r:
                data = json.load(r)
            for upd in data.get("result", []):
                _tg_offset = upd["update_id"] + 1
                msg = upd.get("message", {})
                text = msg.get("text", "").strip()
                if text and str(msg.get("chat", {}).get("id", "")) == TG_CHAT:
                    if text.startswith("/"):
                        threading.Thread(target=handle_tg_command,
                                         args=(text,), daemon=True).start()
                        continue
                    delivered = False
                    with _tg_lock:
                        for jid, state in list(_pending.items()):
                            if not state["event"].is_set():
                                state["reply"] = text
                                state["event"].set()
                                delivered = True
                                break
                    if not delivered:
                        threading.Thread(target=handle_tg_command,
                                         args=("/help " + text,), daemon=True).start()
        except Exception:
            time.sleep(5)

def handle_tg_command(text):
    """Telegram bot commands — Sam's phone remote control."""
    cmd, _, arg = text.partition(" ")
    cmd = cmd.lower()
    arg = arg.strip()
    try:
        if cmd == "/remix":
            if not arg:
                tg("Usage: /remix make a 60s short about CPU cooling from my old clips")
                return
            start_remix(arg)
        elif cmd == "/find":
            if not arg:
                tg("Usage: /find cpu cooling fans")
                return
            hits = search_assets(arg, top=5)
            if not hits:
                tg(f"\U0001f50d Nothing in memory matches “{arg[:60]}”.")
                return
            lines = "\n".join(
                f"• <b>{h.get('topic') or h.get('file','?')}</b> ({h['date'][:10]})\n"
                f"  {h.get('summary','')[:100]}\n"
                f"  \U0001f4e6 {'Drive: ' + h['gdrive_path'] if h.get('archived') else 'still local'}"
                for h in hits)
            tg(f"\U0001f50d <b>Found {len(hits)} match(es):</b>\n{lines}\n\n"
               f"Use /remix <what you want> to build a new edit from these.")
        elif cmd == "/shorts":
            if not arg:
                tg("Usage: /shorts 5 ways to make money with AI")
                return
            jid = uuid.uuid4().hex[:10]
            jobs[jid] = {"status": "running", "progress": 0, "msg": "Queued from Telegram…"}
            threading.Thread(target=run_auto_shorts,
                             args=(jid, arg, {}), daemon=True).start()
            tg(f"⚡ Auto Short started — I'll send the link when it's ready.")
        elif cmd == "/queue":
            q = load_json(PUBLISH_QUEUE)
            q = q if isinstance(q, list) else []
            pend = [x for x in q if x.get("status") == "queued"][-10:]
            if not pend:
                tg("\U0001f4e4 Publish queue is empty.")
            else:
                tg("\U0001f4e4 <b>Publish queue:</b>\n" + "\n".join(
                    f"• {x['channel']} [{x['variant']}] @ {x['publish_at']}" for x in pend) +
                   "\n<i>Auto-publish needs Composio YouTube connected per channel.</i>")
        elif cmd == "/winning":
            tg("\U0001f3c6 Analytics loop needs each channel Composio key connected to "
               "YouTube. Add real keys on the site, then I can pull subs + watch time "
               "and tell you which twin is winning.")
        elif cmd == "/status":
            running = [(k, v) for k, v in jobs.items() if v.get("status") == "running"]
            if running:
                lines = "\n".join(f"• {k[:6]}: {v.get('progress',0)}% — {v.get('msg','')[:60]}"
                                  for k, v in running)
                tg(f"⚙️ <b>{len(running)} job(s) running:</b>\n{lines}")
            else:
                idx_count = len(load_json(ASSET_INDEX) or {})
                tg(f"✅ All quiet. {idx_count} video(s) in memory index.\n"
                   f"Studio: http://192.168.0.78:85")
        else:
            tg("\U0001f916 <b>Mini Studio commands:</b>\n"
               "/remix <request> — new edit from old footage\n"
               "/find <keywords> — search your video memory\n"
               "/shorts <idea> — make a short from scratch\n"
               "/status — what's running\n\n"
               "Plain replies only work while a pipeline is asking you something.")
    except Exception as e:
        tg(f"❌ Command failed: {e}")

threading.Thread(target=tg_poll, daemon=True).start()

def tg_ask(jid, question, timeout=300):
    """Send question to Telegram, wait up to `timeout` seconds for reply. Returns reply str or ''."""
    event = threading.Event()
    with _tg_lock:
        _pending[jid] = {"event": event, "reply": None}
    tg(question)
    got = event.wait(timeout=timeout)
    with _tg_lock:
        reply = _pending.pop(jid, {}).get("reply") or ""
    return reply if got else ""

# ── Excel log ────────────────────────────────────────────────────────────────
EXCEL_COLS = ["Date", "Time", "Job ID", "File", "Channel", "Duration(s)",
              "Transcript Preview", "Scenes", "Clips", "TTS", "Export", "Status",
              "Sam Notes", "Google Drive Path"]

def excel_log(row_data: dict):
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        if EXCEL_LOG.exists():
            wb = load_workbook(str(EXCEL_LOG))
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Pipeline Log"
            ws.append(EXCEL_COLS)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="1a1a2e")
                cell.font = Font(bold=True, color="FFFFFF")
        row = [row_data.get(c, "") for c in EXCEL_COLS]
        ws.append(row)
        # Also write per-channel log if channel folder exists
        ch = row_data.get("Channel", "")
        if ch:
            safe = ch.lower().replace(" ", "_").replace("/", "-")
            ch_log = CHANNELS_ROOT / safe / "log.xlsx"
            ch_log.parent.mkdir(parents=True, exist_ok=True)
            if ch_log.exists():
                wb2 = load_workbook(str(ch_log))
                ws2 = wb2.active
            else:
                wb2 = Workbook(); ws2 = wb2.active
                ws2.title = ch
                ws2.append(EXCEL_COLS)
                for cell in ws2[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1a1a2e")
            ws2.append(row)
            wb2.save(str(ch_log))
        wb.save(str(EXCEL_LOG))
    except Exception as e:
        print(f"[excel] {e}")

# ── Storage helpers ───────────────────────────────────────────────────────────
def load_json(path):
    try:
        return json.loads(path.read_text()) if path.exists() else {}
    except Exception:
        return {}

def save_json(path, data):
    path.write_text(json.dumps(data, indent=2))

jobs = {}

def log(jid, pct, msg, **kw):
    jobs[jid].update({"progress": pct, "msg": msg, **kw})
    print(f"[{jid[:6]}] {pct:3d}% {msg}")

# ── LLM helpers ──────────────────────────────────────────────────────────────
def _openai_call(base_url, api_key, model, messages, max_tokens=2000, timeout=60):
    # Groq blocks Python's default urllib User-Agent — always send a custom one
    import urllib.request
    body = json.dumps({"model": model, "messages": messages,
                       "max_tokens": max_tokens}).encode()
    req = urllib.request.Request(
        f"{base_url}/chat/completions", data=body,
        headers={"Authorization": f"Bearer {api_key}",
                 "Content-Type": "application/json",
                 "User-Agent": "MiniStudio/1.0"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.load(r)["choices"][0]["message"]["content"]

def groq_chat(messages, model=BRAIN_MODEL):
    try:
        return _openai_call(GROQ_BASE, GROQ_KEY, model, messages)
    except Exception:
        if OMNIROUTE_KEY:
            try:
                return _openai_call(OMNIROUTE_BASE, OMNIROUTE_KEY, model, messages)
            except Exception:
                pass
        if GEMINI_KEY:
            return _openai_call(GEMINI_BASE, GEMINI_KEY, "gemini-2.0-flash", messages)
        raise

def groq_vision(image_b64, prompt):
    messages = [{"role": "user", "content": [
        {"type": "text", "text": prompt},
        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_b64}"}}
    ]}]
    for base_url, key, model in [
        (GROQ_BASE, GROQ_KEY, VISION_MODEL),
        (OMNIROUTE_BASE, OMNIROUTE_KEY, "gemini-flash-vision"),
        (GEMINI_BASE, GEMINI_KEY, "gemini-2.0-flash"),
    ]:
        if not key:
            continue
        try:
            return _openai_call(base_url, key, model, messages, max_tokens=300, timeout=30)
        except Exception:
            pass
    return "(vision unavailable)"

# ── Step 1: STT ──────────────────────────────────────────────────────────────
def step_stt(video_path, jid):
    import urllib.request
    with open(video_path, "rb") as f:
        data = f.read()
    boundary = uuid.uuid4().hex
    body = (
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"file\"; "
        f"filename=\"v.mp4\"\r\nContent-Type: video/mp4\r\n\r\n"
    ).encode() + data + f"\r\n--{boundary}--\r\n".encode()
    req = urllib.request.Request(
        f"{WHISPER_URL}/transcribe", data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"})
    with urllib.request.urlopen(req, timeout=600) as r:
        return json.load(r)

# ── Step 2: Frames ───────────────────────────────────────────────────────────
def step_frames(video_path, out_dir, n=8):
    result = subprocess.run(
        ["ffprobe", "-v", "quiet", "-show_entries", "format=duration",
         "-of", "default=noprint_wrappers=1:nokey=1", str(video_path)],
        capture_output=True, text=True)
    dur = float(result.stdout.strip() or "60")
    frames = []
    for i in range(n):
        t = dur * (i + 0.5) / n
        out = Path(out_dir) / f"f{i:02d}.jpg"
        subprocess.run(["ffmpeg", "-y", "-ss", str(t), "-i", str(video_path),
                        "-vframes", "1", "-q:v", "3", str(out)], capture_output=True)
        if out.exists():
            frames.append({"time": round(t, 1), "path": str(out)})
    return frames, dur

# ── Step 3: Vision ───────────────────────────────────────────────────────────
def step_vision(frames):
    scenes = []
    for fr in frames:
        b64 = base64.b64encode(Path(fr["path"]).read_bytes()).decode()
        desc = groq_vision(b64,
            "Describe this video frame in 1-2 sentences. "
            "Include: people, actions, setting, any visible text or products.")
        scenes.append({"time": fr["time"], "description": desc})
    return scenes

# ── Step 4: Brain ────────────────────────────────────────────────────────────
def step_brain(transcript_text, scenes, duration, channels=None, extra_context=""):
    vision_block = "\n".join(f"[{s['time']}s] {s['description']}" for s in scenes)
    ch_hint = channel_hint(channels)
    ctx = f"\nExtra context from creator: {extra_context}" if extra_context else ""
    prompt = f"""Video: {duration:.0f}s long.{ch_hint}{ctx}

TRANSCRIPT:
{transcript_text[:4000]}

SCENE BREAKDOWN:
{vision_block}

Return ONLY valid JSON (no markdown):
{{
  "summary": "2-3 sentences about what this video covers",
  "topic": "one-line topic/niche",
  "best_clips": [
    {{"start": 0.0, "end": 30.0, "title": "clip title", "reason": "why this works"}}
  ],
  "hook": "grabby opening line for TikTok/YouTube Shorts",
  "script": "60-90 second spoken script based on this content, written naturally for TTS",
  "tags": ["tag1","tag2","tag3","tag4","tag5"]
}}"""
    raw = groq_chat([
        {"role": "system", "content": "You are a video editor AI. Return only valid JSON."},
        {"role": "user", "content": prompt}
    ])
    raw = re.sub(r"^```json?\s*", "", raw.strip())
    raw = re.sub(r"\s*```$", "", raw.strip())
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return {"summary": raw[:500], "best_clips": [], "hook": "", "script": raw[:500], "tags": []}

# ── Step 4b: Shorts brain ────────────────────────────────────────────────────
def brain_shorts(idea, channels=None):
    ch_hint = channel_hint(channels)
    prompt = f"""Create a TikTok/YouTube Short from this idea: {idea}{ch_hint}

Return ONLY valid JSON:
{{
  "title": "Short title",
  "hook": "Grabby first line (spoken in first 3 seconds)",
  "script": "Complete 45-60 second spoken script — conversational, punchy, no stage directions",
  "thumbnail_prompt": "Visual description for AI image generation",
  "tags": ["tag1","tag2","tag3","tag4","tag5"],
  "topic": "niche/topic"
}}"""
    raw = groq_chat([
        {"role": "system", "content": "You are a viral short-form video creator. Return only valid JSON."},
        {"role": "user", "content": prompt}
    ])
    raw = re.sub(r"^```json?\s*", "", raw.strip())
    raw = re.sub(r"\s*```$", "", raw.strip())
    try:
        return json.loads(raw)
    except Exception:
        return {"title": idea[:60], "hook": "", "script": idea,
                "thumbnail_prompt": idea, "tags": [], "topic": ""}

# ── Step 5: Slice ────────────────────────────────────────────────────────────
def step_slice(video_path, clips, jid):
    out_clips = []
    for i, clip in enumerate(clips[:4]):
        s, e = float(clip.get("start", 0)), float(clip.get("end", 30))
        out = EXPORT / f"{jid}_clip{i+1}.mp4"
        subprocess.run([
            "ffmpeg", "-y", "-i", str(video_path),
            "-ss", str(s), "-to", str(e),
            "-vf", "scale=1920:1080:force_original_aspect_ratio=decrease,pad=1920:1080:(ow-iw)/2:(oh-ih)/2",
            "-c:v", "libx264", "-profile:v", "high", "-level", "4.0", "-pix_fmt", "yuv420p",
            "-c:a", "aac", "-b:a", "128k", "-preset", "fast", "-crf", "23", str(out)
        ], capture_output=True)
        if out.exists():
            out_clips.append({"file": out.name, "title": clip.get("title", f"Clip {i+1}"),
                              "start": s, "end": e})
    return out_clips

# ── Step 6: TTS ──────────────────────────────────────────────────────────────
def gtts_wav(text):
    """Google free TTS — demo voice until Sam adds his own model."""
    from gtts import gTTS
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        mp3 = Path(td) / "t.mp3"
        wav = Path(td) / "t.wav"
        gTTS(text=text[:2500], lang="en", tld="co.uk").save(str(mp3))
        subprocess.run(["ffmpeg", "-y", "-i", str(mp3), "-ar", "24000", str(wav)],
                       capture_output=True, check=True)
        return wav.read_bytes()

def step_tts(script, voice_path=None, language="en"):
    # No cloned voice chosen → Google free TTS (clearer than generic XTTS)
    if not voice_path:
        try:
            return gtts_wav(script)
        except Exception as e:
            print(f"[gtts] failed ({e}), falling back to XTTS")
    import urllib.request
    body = {"text": script[:2500], "language": language or "en"}
    if voice_path:
        body["speaker_wav"] = voice_path
    req = urllib.request.Request(
        f"{TTS_URL}/generate", data=json.dumps(body).encode(),
        headers={"Content-Type": "application/json"})
    # XTTS on CPU runs ~2x real-time — a 90s script can take several minutes
    with urllib.request.urlopen(req, timeout=900) as r:
        return r.read()

# ── Auto Shorts video assembly ───────────────────────────────────────────────
def assemble_shorts_video(image_path, audio_path, out_path):
    result = subprocess.run(
        ["ffprobe", "-v", "quiet", "-show_entries", "format=duration",
         "-of", "default=noprint_wrappers=1:nokey=1", str(audio_path)],
        capture_output=True, text=True)
    dur = float(result.stdout.strip() or "60")
    subprocess.run([
        "ffmpeg", "-y",
        "-loop", "1", "-i", str(image_path),
        "-i", str(audio_path),
        "-c:v", "libx264", "-pix_fmt", "yuv420p", "-tune", "stillimage",
        "-c:a", "aac", "-b:a", "128k",
        "-vf", "scale=1080:1920:force_original_aspect_ratio=decrease,pad=1080:1920:-1:-1:color=black",
        "-shortest", "-t", str(dur + 0.5), "-preset", "fast", "-crf", "23", str(out_path)
    ], capture_output=True, check=True)

# ── Step 7: Channel export ───────────────────────────────────────────────────
def step_channel_export(video_path, voice_wav, clone_voice=None, clone_text=None,
                        export_channels=None):
    import urllib.request
    payload = {
        "video_path": str(video_path),
        "voice_path": str(voice_wav) if voice_wav else None,
        "clone_voice": clone_voice, "clone_text": clone_text,
        "channels": export_channels or ["A", "B", "C"],
        "bg_volume": 0.12, "voice_volume": 1.0, "clone_volume": 1.0,
    }
    req = urllib.request.Request(
        f"{CHANNEL_URL}/export", data=json.dumps(payload).encode(),
        headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.load(r)
    except Exception as e:
        return {"error": str(e)}

# ── Asset packages + Hermes memory index ─────────────────────────────────────
def build_asset_package(jid, video_path, brain, transcript, segments, scenes,
                        clips, tts_file, ch_names, sam_notes="",
                        priority="normal", deadline="", raw_gdrive=""):
    """Bundle everything about a finished job into one folder + index record.
    The package folder is what gets archived to Google Drive; the index stays
    local forever and is Hermes' searchable memory."""
    slug = re.sub(r"[^a-z0-9]+", "_",
                  (brain.get("topic") or Path(video_path).stem).lower())[:40].strip("_") or "video"
    pkg_name = f"{jid}_{slug}"
    pkg_dir = PACKAGES_DIR / pkg_name
    pkg_dir.mkdir(parents=True, exist_ok=True)
    try:
        shutil.copy2(video_path, pkg_dir / "raw_source.mp4")
    except Exception as e:
        print(f"[package] raw copy failed: {e}")
    for c in clips:
        src = EXPORT / c["file"]
        if src.exists():
            try:
                shutil.copy2(src, pkg_dir / c["file"])
            except Exception:
                pass
    if tts_file and Path(tts_file).exists():
        try:
            shutil.copy2(tts_file, pkg_dir / "voiceover.wav")
        except Exception:
            pass
    (pkg_dir / "vision_scan.json").write_text(json.dumps(scenes, indent=2))
    (pkg_dir / "script_and_timestamps.json").write_text(json.dumps({
        "transcript": transcript, "segments": segments,
        "script": brain.get("script", ""), "cuts": brain.get("best_clips", []),
        "hook": brain.get("hook", ""), "tags": brain.get("tags", []),
        "topic": brain.get("topic", ""), "summary": brain.get("summary", ""),
    }, indent=2))

    gdrive_path = f"{GDRIVE_ARCHIVE}/{time.strftime('%Y-%m')}/{pkg_name}"
    record = {
        "jid": jid, "package": pkg_name, "file": Path(video_path).name,
        "date": time.strftime("%Y-%m-%d %H:%M"), "channel": ch_names,
        "topic": brain.get("topic", ""), "summary": brain.get("summary", ""),
        "script": brain.get("script", ""), "hook": brain.get("hook", ""),
        "tags": brain.get("tags", []),
        "transcript": transcript,
        "segments": segments,
        "scenes_text": " ".join(s.get("description", "") for s in scenes),
        "sam_notes": sam_notes,
        "gdrive_path": gdrive_path,
        "archived": False,
        "location": "local",
        "priority": priority,
        "deadline": deadline,
        "raw_gdrive": raw_gdrive,
    }
    idx = load_json(ASSET_INDEX) or {}
    idx[jid] = record
    save_json(ASSET_INDEX, idx)
    return gdrive_path

def search_assets(query, top=3):
    """Keyword search across Hermes' memory index. Fast — text only."""
    words = [w for w in re.findall(r"[a-z0-9]+", query.lower()) if len(w) > 2]
    if not words:
        return []
    idx = load_json(ASSET_INDEX) or {}
    scored = []
    for rec in idx.values():
        haystack = " ".join([
            rec.get("topic", ""), rec.get("summary", ""), rec.get("script", ""),
            rec.get("transcript", ""), rec.get("scenes_text", ""),
            " ".join(rec.get("tags", [])), rec.get("sam_notes", ""),
            rec.get("file", ""),
        ]).lower()
        score = sum(haystack.count(w) for w in words)
        if score > 0:
            scored.append((score, rec))
    scored.sort(key=lambda x: -x[0])
    return [r for _, r in scored[:top]]

PC88_ARCHIVE = "pc88:/srv/backup-staging/studio-archive"

def ensure_asset_local(rec):
    """Return a local path to the asset's raw source. 3-tier recall:
    CT500 local → 0.88 warm storage → Google Drive."""
    local = PACKAGES_DIR / rec["package"] / "raw_source.mp4"
    if local.exists():
        return local
    recall = RECALL_DIR / rec["package"] / "raw_source.mp4"
    if recall.exists():
        return recall
    check = subprocess.run(["rclone", "listremotes"], capture_output=True, text=True)
    recall.parent.mkdir(parents=True, exist_ok=True)
    # Try 0.88 first (fast LAN pull), then Google Drive
    candidates = []
    if "pc88:" in check.stdout:
        candidates.append(f"{PC88_ARCHIVE}/{rec['package']}/raw_source.mp4")
    if "gdrive:" in check.stdout:
        candidates.append(f"gdrive:{rec['gdrive_path']}/raw_source.mp4")
        if rec.get("raw_gdrive"):
            candidates.append(f"gdrive:{rec['raw_gdrive']}")
    for src in candidates:
        subprocess.run(["rclone", "copyto", src, str(recall)],
                       capture_output=True, timeout=1800)
        if recall.exists():
            return recall
    return None

def channel_hint(channels):
    """Describe target channels (style/format/voice) for the AI brain."""
    if not channels:
        return ""
    parts = []
    for c in channels:
        bits = [c.get("name", "")]
        if c.get("style"):  bits.append(f"style: {c['style']}")
        if c.get("format"): bits.append(f"format: {c['format']} video")
        if c.get("niche"):  bits.append(f"niche: {c['niche']}")
        if c.get("linked_notes"): bits.append(f"twin needs: {c['linked_notes']}")
        parts.append(" — ".join(bits))
    return "\nTarget channels: " + "; ".join(parts)

# ── Helpers ───────────────────────────────────────────────────────────────────
def resolve_channels(channel_ids):
    """Resolve ids to channel objects, auto-including linked twin channels."""
    if not channel_ids:
        return []
    chs = load_json(CHANNELS_FILE)
    out, seen = [], set()
    def add(cid):
        if cid in seen or cid not in chs:
            return
        seen.add(cid)
        out.append(chs[cid])
        link = chs[cid].get("linked_channel", "")
        if link:
            add(link)
    for cid in channel_ids:
        add(cid)
    return out

def distribute_exports(jid, channels, base_time=None):
    """Render once, deliver per-twin: copy each channel's audio variant into its
    exports folder and queue staggered publishes (Composio posts when wired)."""
    base_time = base_time or time.time()
    exports_src = Path("/opt/studio/media/exports")
    delivered = []
    for attempt in range(60):   # wait up to 10 min for the A/B/C renders
        fresh = {f.name.rsplit("_ch", 1)[1][0]: f
                 for f in exports_src.glob("*_ch[ABC].mp4")
                 if f.stat().st_mtime > base_time - 60}
        if fresh:
            time.sleep(10)
            fresh = {f.name.rsplit("_ch", 1)[1][0]: f
                     for f in exports_src.glob("*_ch[ABC].mp4")
                     if f.stat().st_mtime > base_time - 60}
            break
        time.sleep(10)
    else:
        return []
    queue = load_json(PUBLISH_QUEUE) or []
    if isinstance(queue, dict):
        queue = []
    offset = 0.0
    for ch in channels:
        var = ch.get("audio_variant", "C")
        src_f = fresh.get(var)
        if not src_f:
            continue
        safe = ch.get("folder_name") or ch["name"].lower().replace(" ", "_")
        dest_dir = CHANNELS_ROOT / safe / "exports"
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest = dest_dir / f"{jid}_{var}.mp4"
        shutil.copy2(src_f, dest)
        stagger = ch.get("stagger_hours") or offset
        pub_at = time.time() + stagger * 3600
        queue.append({"jid": jid, "channel": ch["name"], "file": str(dest),
                      "variant": var,
                      "publish_at": time.strftime("%Y-%m-%d %H:%M",
                                                  time.localtime(pub_at)),
                      "composio": ch.get("composio_connection", ""),
                      "status": "queued"})
        delivered.append(f"{ch['name']} gets variant {var}, publish +{stagger:g}h")
        offset += 4.0   # default twin stagger
    Path(PUBLISH_QUEUE).write_text(json.dumps(queue, indent=2))
    return delivered

def _channel_voice(channels):
    for ch in channels:
        if ch.get("voice_path"):
            return ch["voice_path"]
    return None

def _save_blueprint_auto(options, brain, channels):
    bps = load_json(BLUEPRINTS_FILE)
    bid = uuid.uuid4().hex[:8]
    bps[bid] = {
        "id": bid, "name": options["save_blueprint_name"],
        "voice_path": options.get("voice_path", ""),
        "channel_ids": options.get("channel_ids", []),
        "topic": brain.get("topic", ""),
        "created": time.strftime("%Y-%m-%d %H:%M"),
    }
    save_json(BLUEPRINTS_FILE, bps)

# ── Main pipeline ─────────────────────────────────────────────────────────────
def run_pipeline(jid, video_path, options):
    try:
        wdir = WORK / jid; wdir.mkdir(exist_ok=True)
        channels = resolve_channels(options.get("channel_ids", []))
        fname = Path(video_path).name
        ch_names = ", ".join(c.get("name", "") for c in channels) or "none"

        # ── STEP 0: raw video → Google Drive FIRST (safety, in background) ───
        raw_gdrive = f"mini-studio/raw/{time.strftime('%Y-%m')}/{fname}"
        def _backup_raw():
            r = subprocess.run(["rclone", "copyto", str(video_path),
                                f"gdrive:{raw_gdrive}"],
                               capture_output=True, timeout=3600)
            if r.returncode == 0:
                print(f"[{jid[:6]}] raw backed up to Drive: {raw_gdrive}")
                tg(f"☁️ <b>{fname}</b> safe on Google Drive — "
                   f"source card/file can be wiped.")
            else:
                print(f"[{jid[:6]}] raw Drive backup FAILED: {r.stderr.decode()[:150]}")
                queue_upload(video_path, raw_gdrive)
                tg(f"\U0001f4f5 <b>{fname}</b>: no internet for Drive backup — "
                   f"queued, will auto-upload when back online. "
                   f"Do NOT format the card until I confirm!")
        threading.Thread(target=_backup_raw, daemon=True).start()
        options["raw_gdrive"] = raw_gdrive

        # ── STEP 1: Transcribe ───────────────────────────────────────────────
        log(jid, 5, "Step 1 — Transcribing (Whisper)…")
        try:
            stt = step_stt(video_path, jid)
            transcript = stt.get("text", "")
            segments   = stt.get("segments", [])
        except Exception as e:
            transcript = ""; segments = []
            log(jid, 10, f"STT failed ({e}), continuing with vision only…")

        # ── STEP 2: Frames ───────────────────────────────────────────────────
        log(jid, 20, "Step 2 — Extracting frames…")
        frames, duration = step_frames(video_path, wdir)

        # ── STEP 3: Vision ───────────────────────────────────────────────────
        log(jid, 30, f"Step 3 — Vision ({len(frames)} frames)…")
        scenes = []
        for i, fr in enumerate(frames):
            log(jid, 30 + i * 4, f"Vision frame {i+1}/{len(frames)}…")
            b64 = base64.b64encode(Path(fr["path"]).read_bytes()).decode()
            desc = groq_vision(b64,
                "Describe this video frame in 1-2 sentences. "
                "Include people, actions, setting, visible text or products.")
            scenes.append({"time": fr["time"], "description": desc})

        # ── Telegram: send scan summary, ask Sam for context ─────────────────
        preview_transcript = transcript[:300].replace("\n", " ") if transcript else "(no speech)"
        preview_scenes = " | ".join(s["description"][:60] for s in scenes[:3])
        sam_question = (
            f"\U0001f50d <b>Scan complete!</b>\n"
            f"\U0001f3ac <b>{fname}</b>  ({duration:.0f}s)\n\n"
            f"\U0001f4dd <b>Transcript:</b> {preview_transcript}\n\n"
            f"\U0001f3a5 <b>Scenes:</b> {preview_scenes}\n\n"
            f"\U0001f4e1 Channels: {ch_names}\n\n"
            f"<b>What do we do?</b> Tell me e.g. \"make it a quick short about the CPU tip\" "
            f"— or reply <b>ok</b> and I'll pick the best cuts myself.\n"
            f"Say <b>daily</b>/<b>urgent</b> to keep it hot on the studio, "
            f"or <b>weekly</b>/<b>low priority</b> to park it on 0.88.\n"
            f"<i>(Auto-continues in 5 min)</i>"
        )
        full_auto = bool(channels) and all(c.get("auto_mode") for c in channels)
        if full_auto:
            log(jid, 62, "Channel is in Auto Mode — skipping interview…")
            tg(f"\U0001f916 <b>{fname}</b>: channel in Auto Mode — building with "
               f"its saved style/voice, no questions asked.")
            sam_reply = ""
        else:
            log(jid, 62, "Waiting for your Telegram reply (or auto-continue in 5 min)…")
            sam_reply = tg_ask(jid, sam_question, timeout=300)
        extra_context = sam_reply if sam_reply.lower() not in ("ok", "okay", "") else ""
        if sam_reply:
            log(jid, 63, f"Got your reply: {sam_reply[:60]}")
        else:
            log(jid, 63, "No reply — continuing automatically…")

        # Priority keywords in the reply override the upload setting
        low_reply = sam_reply.lower()
        if any(k in low_reply for k in ("daily", "urgent", "high priority", "asap", "today")):
            options["priority"] = "high"
        elif any(k in low_reply for k in ("low priority", "weekly", "no rush", "whenever")):
            options["priority"] = "low"

        # ── STEP 4: Brain ────────────────────────────────────────────────────
        log(jid, 65, "Step 4 — Groq Brain…")
        brain = step_brain(transcript, scenes, duration, channels, extra_context)

        # ── Telegram: send the timeline plan back to Sam ─────────────────────
        plan_clips = brain.get("best_clips", [])
        plan_lines = "\n".join(
            f"  {i+1}. [{c.get('start',0):.0f}s–{c.get('end',0):.0f}s] {c.get('title','')}"
            for i, c in enumerate(plan_clips[:4])
        ) or "  (no clips suggested)"
        tg(
            f"\U0001f4cb <b>Timeline plan ready</b>\n"
            f"\U0001f3ac {fname}\n\n"
            f"✂️ <b>Cuts:</b>\n{plan_lines}\n\n"
            f"\U0001f3a4 <b>Hook:</b> {brain.get('hook','')[:100]}\n\n"
            f"\U0001f4dc <b>Script:</b> {brain.get('script','')[:400]}\n\n"
            f"Now slicing + generating voice… I'll message when export is done."
        )

        # ── Save draft for OpenCut ───────────────────────────────────────────
        draft = {
            "job_id": jid, "file": fname, "video_path": str(video_path),
            "created": time.strftime("%Y-%m-%d %H:%M"),
            "duration": duration, "channel": ch_names,
            "summary": brain.get("summary", ""), "topic": brain.get("topic", ""),
            "hook": brain.get("hook", ""), "script": brain.get("script", ""),
            "tags": brain.get("tags", []),
            "cuts": plan_clips, "sam_notes": extra_context,
            "transcript": transcript, "segments": segments, "scenes": scenes,
            "status": "pending",
        }
        save_json(DRAFTS_DIR / f"{jid}.json", draft)

        # ── STEP 5: Slice ────────────────────────────────────────────────────
        log(jid, 75, "Step 5 — FFmpeg smart slice…")
        clips = step_slice(video_path, brain.get("best_clips", []), jid)

        # ── STEP 6: TTS ──────────────────────────────────────────────────────
        log(jid, 83, "Step 6 — TTS voice…")
        script = brain.get("script", transcript[:800])
        tts_file = None
        voice_path = options.get("voice_path") or _channel_voice(channels)
        lang = next((c.get("language") for c in channels if c.get("language")), "en")
        try:
            wav = step_tts(script, voice_path, language=lang)
            tts_file = str(EXPORT / f"{jid}_script.wav")
            Path(tts_file).write_bytes(wav)
        except Exception as e:
            log(jid, 84, f"TTS failed ({e})")

        # ── STEP 7: Channel export ───────────────────────────────────────────
        log(jid, 90, "Step 7 — Channel export A/B/C…")
        channel_job = None
        if tts_file and clips:
            channel_job = step_channel_export(
                clips[0]["file"] if clips else video_path,
                tts_file, clone_voice=voice_path, clone_text=script)
            if channels:
                def _distribute():
                    d = distribute_exports(jid, channels)
                    if d:
                        tg("\U0001f4e4 <b>Twin delivery queued:</b>\n" + "\n".join(d) +
                           "\n(auto-publish activates when Composio YouTube is connected)")
                threading.Thread(target=_distribute, daemon=True).start()

        if options.get("save_blueprint_name"):
            _save_blueprint_auto(options, brain, channels)

        # Update draft with final artefacts
        draft.update({"sliced_clips": clips, "tts_file": tts_file, "status": "ready"})
        save_json(DRAFTS_DIR / f"{jid}.json", draft)

        # ── Format renders: vertical 9:16 short when the channel wants it ────
        vertical_file = None
        formats = {c.get("format", "both") for c in channels} or {"both"}
        if formats & {"short", "both"} and clips:
            log(jid, 93, "Rendering 9:16 short version…")
            src_clip = EXPORT / clips[0]["file"]
            vout = EXPORT / f"{jid}_short916.mp4"
            subprocess.run([
                "ffmpeg", "-y", "-i", str(src_clip),
                "-vf", "crop=ih*9/16:ih,scale=1080:1920",
                "-c:v", "libx264", "-pix_fmt", "yuv420p", "-c:a", "aac", "-preset", "fast", "-crf", "23",
                str(vout)], capture_output=True)
            if vout.exists() and vout.stat().st_size > 0:
                vertical_file = vout.name

        result = {
            "duration": duration, "transcript": transcript, "segments": segments,
            "scenes": scenes, "summary": brain.get("summary", ""),
            "topic": brain.get("topic", ""), "hook": brain.get("hook", ""),
            "script": script, "tags": brain.get("tags", []),
            "best_clips": clips, "tts_file": tts_file,
            "vertical_short": vertical_file,
            "channel_export": channel_job,
            "channels_used": [c.get("name") for c in channels],
            "export_dir": str(EXPORT),
        }

        # ── Asset package + Hermes memory index ─────────────────────────────
        gdrive_path = ""
        try:
            gdrive_path = build_asset_package(
                jid, video_path, brain, transcript, segments, scenes,
                clips, tts_file, ch_names, sam_notes=extra_context,
                priority=options.get("priority", "normal"),
                deadline=options.get("deadline", ""),
                raw_gdrive=options.get("raw_gdrive", ""))
            # Raw lives in the package + on Drive — drop the duplicate upload
            try:
                if Path(video_path).parent == UPLOAD:
                    Path(video_path).unlink(missing_ok=True)
            except Exception:
                pass
        except Exception as e:
            print(f"[package] failed: {e}")

        # ── Excel log ────────────────────────────────────────────────────────
        excel_log({
            "Date": time.strftime("%Y-%m-%d"), "Time": time.strftime("%H:%M"),
            "Job ID": jid, "File": fname, "Channel": ch_names,
            "Duration(s)": round(duration, 1),
            "Transcript Preview": transcript[:150],
            "Scenes": len(scenes), "Clips": len(clips),
            "TTS": "yes" if tts_file else "no",
            "Export": "yes" if channel_job and "error" not in channel_job else "no",
            "Status": "done", "Sam Notes": extra_context[:200],
            "Google Drive Path": gdrive_path,
        })

        # ── Done notification ────────────────────────────────────────────────
        tg(
            f"✅ <b>Pipeline done!</b>\n"
            f"\U0001f3ac {fname}\n"
            f"\U0001f4dd {brain.get('summary','')[:120]}\n"
            f"✂️ {len(clips)} clips | {ch_names}\n"
            f"\U0001f3b5 TTS: {'yes' if tts_file else 'no'} | Export A/B/C: {'yes' if channel_job and 'error' not in channel_job else 'no'}"
            + (" | \U0001f4f1 9:16 short: yes" if vertical_file else "") + "\n\n"
            f"\U0001f39e <b>Open on OpenCut timeline (cuts + captions + voice ready):</b>\n"
            f"http://192.168.0.78:9500/ai-draft/{jid}"
        )
        log(jid, 100, "✓ Pipeline complete!", status="done", result=result)

    except Exception as e:
        import traceback
        tg(f"❌ <b>Pipeline error</b>\n{e}")
        log(jid, 0, f"Error: {e}", status="error", trace=traceback.format_exc())

# ── Auto Shorts pipeline ──────────────────────────────────────────────────────
def run_auto_shorts(jid, idea, options):
    try:
        wdir = WORK / jid; wdir.mkdir(exist_ok=True)
        channels = resolve_channels(options.get("channel_ids", []))
        ch_names = ", ".join(c.get("name", "") for c in channels) or "none"

        tg(f"⚡ <b>Auto Short started</b>\n\U0001f4a1 {idea[:80]}\n\U0001f4e1 {ch_names}")

        log(jid, 10, "Generating script…")
        brain = brain_shorts(idea, channels)
        script = brain.get("script", idea)

        log(jid, 30, "Generating voice (XTTS)…")
        voice_path = options.get("voice_path") or _channel_voice(channels)
        tts_file = None
        try:
            wav = step_tts(script, voice_path)
            tts_file = str(WORK / jid / "speech.wav")
            Path(tts_file).write_bytes(wav)
        except Exception as e:
            log(jid, 35, f"TTS failed ({e})")

        log(jid, 50, "Generating thumbnail (Pollinations)…")
        img_file = None
        try:
            import urllib.request as _ur
            prompt = brain.get("thumbnail_prompt", idea)
            body = json.dumps({"prompt": prompt, "width": 1080, "height": 1920,
                               "model": "flux"}).encode()
            req = _ur.Request("http://localhost:8423/generate", data=body,
                              headers={"Content-Type": "application/json"})
            with _ur.urlopen(req, timeout=60) as r:
                img_data = r.read()
            img_file = str(WORK / jid / "thumb.jpg")
            Path(img_file).write_bytes(img_data)
        except Exception as e:
            log(jid, 55, f"Image gen failed ({e})")

        log(jid, 65, "Assembling video…")
        final_vid = str(EXPORT / f"{jid}_short.mp4")
        if tts_file and img_file and Path(img_file).exists():
            assemble_shorts_video(img_file, tts_file, final_vid)
        elif tts_file:
            subprocess.run([
                "ffmpeg", "-y", "-f", "lavfi",
                "-i", "color=c=black:s=1080x1920:r=25",
                "-i", str(tts_file), "-c:v", "libx264", "-pix_fmt", "yuv420p",
                "-c:a", "aac", "-shortest", final_vid
            ], capture_output=True)

        log(jid, 85, "Channel export A/B/C…")
        channel_job = None
        if tts_file and Path(final_vid).exists() and Path(final_vid).stat().st_size > 0:
            channel_job = step_channel_export(final_vid, tts_file,
                                              clone_voice=voice_path, clone_text=script)

        if options.get("save_blueprint_name"):
            _save_blueprint_auto(options, brain, channels)

        result = {
            "title": brain.get("title", idea[:60]),
            "hook": brain.get("hook", ""), "script": script,
            "tags": brain.get("tags", []), "topic": brain.get("topic", ""),
            "video": Path(final_vid).name if Path(final_vid).exists() else None,
            "tts_file": tts_file, "thumbnail": img_file,
            "channel_export": channel_job,
            "channels_used": [c.get("name") for c in channels],
        }

        excel_log({
            "Date": time.strftime("%Y-%m-%d"), "Time": time.strftime("%H:%M"),
            "Job ID": jid, "File": f"[AUTO] {brain.get('title',idea[:40])}",
            "Channel": ch_names, "Duration(s)": "",
            "Transcript Preview": script[:150], "Scenes": 0, "Clips": 1,
            "TTS": "yes" if tts_file else "no",
            "Export": "yes" if channel_job and "error" not in channel_job else "no",
            "Status": "done", "Sam Notes": idea[:200],
        })

        tg(
            f"⚡ <b>Auto Short ready!</b>\n"
            f"\U0001f4a1 {brain.get('title','')[:60]}\n"
            f"\U0001f3b5 {brain.get('hook','')[:80]}\n"
            f"\U0001f4e1 {ch_names}\n"
            f"http://192.168.0.78:9530"
        )
        log(jid, 100, "✓ Auto Short ready!", status="done", result=result)

    except Exception as e:
        import traceback
        tg(f"❌ <b>Auto Shorts error</b>\n{e}")
        log(jid, 0, f"Error: {e}", status="error", trace=traceback.format_exc())

# ── Projects: multi-file video builds ─────────────────────────────────────────
def project_dir(proj):
    return Path(proj.get("dir") or (PROJECTS_ROOT / proj["slug"]))

def project_counts(proj):
    base = project_dir(proj) if isinstance(proj, dict) else PROJECTS_ROOT / proj
    return {sub: len([f for f in (base / sub).iterdir() if f.is_file()])
            if (base / sub).is_dir() else 0 for sub in PROJECT_SUBS}

def run_project_check(jid, pid):
    try:
        projects = load_json(PROJECTS_FILE)
        proj = projects.get(pid)
        if not proj:
            log(jid, 0, "Project not found", status="error"); return
        slug = proj["slug"]; pdir = project_dir(proj)
        channels = resolve_channels(proj.get("channel_ids", []))
        ch_names = ", ".join(c.get("name", "") for c in channels) or "none"
        voice_mode = proj.get("voice_mode", "both")
        counts = project_counts(proj)

        a_rolls = sorted(f for f in (pdir / "a-roll").iterdir()
                         if f.suffix.lower() in VIDEO_EXTS)
        b_rolls = sorted(f for f in (pdir / "b-roll").iterdir()
                         if f.suffix.lower() in VIDEO_EXTS)
        pics    = sorted(f for f in (pdir / "pics").iterdir() if f.is_file())
        if not a_rolls and not b_rolls:
            tg(f"❌ <b>{proj['name']}</b>: no videos in a-roll or b-roll yet.")
            log(jid, 0, "No videos in project", status="error"); return

        tg(f"\U0001f3ac <b>AI Check started: {proj['name']}</b>\n"
           f"\U0001f4c1 {counts['a-roll']} a-roll, {counts['b-roll']} b-roll, "
           f"{counts['pics']} pics, {counts['other']} other\n"
           f"\U0001f4e1 {ch_names} | voice: {voice_mode} | "
           f"by: {proj.get('deadline') or 'no deadline'}\n"
           f"☁️ Backing up to Drive + scanning…")

        # Drive FIRST — whole project folder in background
        raw_gdrive = f"mini-studio/raw/projects/{slug}"
        def _backup():
            r = subprocess.run(["rclone", "copy", str(pdir), f"gdrive:{raw_gdrive}"],
                               capture_output=True, timeout=7200)
            tg(f"☁️ <b>{proj['name']}</b> fully backed up to Drive — cards can be wiped."
               if r.returncode == 0 else
               f"⚠️ Drive backup FAILED for {proj['name']} — do NOT wipe cards!")
        threading.Thread(target=_backup, daemon=True).start()

        # Scan every clip
        log(jid, 10, f"Scanning {len(a_rolls)} a-roll + {len(b_rolls)} b-roll clips…")
        clip_info = []
        total = len(a_rolls) + len(b_rolls)
        done = 0
        for role, files in (("a-roll", a_rolls), ("b-roll", b_rolls)):
            for f in files:
                done += 1
                log(jid, 10 + int(40 * done / max(total, 1)),
                    f"Scanning {role}: {f.name} ({done}/{total})…")
                transcript, segments = "", []
                if role == "a-roll":
                    try:
                        stt = step_stt(f, jid)
                        transcript, segments = stt.get("text", ""), stt.get("segments", [])
                    except Exception as e:
                        print(f"[proj] stt {f.name}: {e}")
                wdir = WORK / jid / f.stem; wdir.mkdir(parents=True, exist_ok=True)
                frames, dur = step_frames(f, wdir, n=4 if role == "b-roll" else 6)
                scenes = step_vision(frames)
                clip_info.append({"file": f.name, "path": str(f), "role": role,
                                  "duration": dur, "transcript": transcript,
                                  "segments": segments, "scenes": scenes})

        # Telegram: summary + ask Sam
        summary_lines = "\n".join(
            f"• [{c['role']}] {c['file']} ({c['duration']:.0f}s): "
            f"{(c['transcript'][:60] or c['scenes'][0]['description'][:60] if c['scenes'] else '')}"
            for c in clip_info[:10])
        # Director Interview — Hermes pitches an edit based on the brief
        bp_name = ""
        if proj.get("blueprint_id"):
            bp_name = (load_json(BLUEPRINTS_FILE).get(proj["blueprint_id"]) or {}).get("name", "")
        interview = ""
        try:
            interview = groq_chat([
                {"role": "system", "content":
                 "You are Hermes, an executive producer texting the creator on Telegram. "
                 "2-4 sentences max: pitch how you'd cut this video, then ask ONE "
                 "sharp question about the direction. Friendly, no lists, no markdown."},
                {"role": "user", "content":
                 f"Project: {proj['name']} | video: {proj.get('video_name','')}\n"
                 f"Creator's idea: {proj.get('bio') or proj.get('notes') or 'not given'}\n"
                 f"Blueprint: {bp_name or 'none'} | voice: {voice_mode} | "
                 f"deadline: {proj.get('deadline') or 'none'}\n"
                 f"Footage scanned:\n{summary_lines}"}])[:600]
        except Exception:
            pass
        if not interview:
            interview = "What do we do? Tell me the story/angle, or ok to let me build it."
        log(jid, 55, "Waiting for your Telegram reply (5 min)…")
        sam_reply = tg_ask(jid,
            f"\U0001f3ac <b>{proj['name']} — scan done!</b>\n"
            f"\U0001f4c1 {counts['a-roll']} a-roll · {counts['b-roll']} b-roll · "
            f"{counts['pics']} pics\n{summary_lines}\n\n"
            f"\U0001f9e0 {interview}\n<i>(reply, or 'ok' — auto-continues in 5 min)</i>",
            timeout=300)
        extra = sam_reply if sam_reply.lower() not in ("ok", "okay", "") else ""

        # Brain across all clips
        log(jid, 60, "Groq Brain: building the video plan…")
        blocks = []
        for c in clip_info:
            segs = "\n".join(f"    [{s['start']:.1f}-{s['end']:.1f}s] {s['text'].strip()}"
                             for s in (c["segments"] or [])[:30])
            vis = " | ".join(s["description"][:60] for s in c["scenes"][:3])
            blocks.append(f"CLIP {c['file']} ({c['role']}, {c['duration']:.0f}s)\n"
                          f"  visuals: {vis}\n  speech:\n{segs or '    (none)'}")
        pic_names = ", ".join(p.name for p in pics[:15]) or "none"
        prompt = f"""Project: {proj['name']} | video name: {proj.get('video_name','')}
Channels: {ch_names}{channel_hint(channels)} | creator notes: {extra or proj.get('notes','') or 'none'}
Available pictures: {pic_names}

CLIPS:
{chr(10).join(blocks)}

Build the video. Return ONLY valid JSON:
{{
  "title": "...", "hook": "...",
  "script": "45-90s voiceover script, natural for TTS",
  "clips": [{{"file": "<clip filename>", "start": 0.0, "end": 8.0, "title": "why"}}],
  "tags": ["t1","t2","t3"]
}}
Rules: 2-8 cuts, 3-25s each, a-roll for speech moments, b-roll for visuals."""
        raw = groq_chat([
            {"role": "system", "content": "You are a video editor AI. ONLY JSON."},
            {"role": "user", "content": prompt}])
        raw = re.sub(r"^```json?\s*|\s*```$", "", raw.strip())
        m = re.search(r"\{.*\}", raw, re.S)
        brain = json.loads(m.group(0) if m else raw)
        script = brain.get("script", "")

        # Slice + concat
        log(jid, 72, "Slicing + assembling…")
        by_name = {c["file"]: c["path"] for c in clip_info}
        sliced, cuts, cursor = [], [], 0.0
        for i, c in enumerate(brain.get("clips", [])[:8]):
            src = by_name.get(c.get("file", ""))
            if not src: continue
            s, e = float(c.get("start", 0)), float(c.get("end", 0))
            if e - s < 1: continue
            out = EXPORT / f"{jid}_clip{i+1}.mp4"
            subprocess.run(["ffmpeg", "-y", "-i", src, "-ss", str(s), "-to", str(e),
                "-vf", "scale=1280:720:force_original_aspect_ratio=decrease,"
                       "pad=1280:720:-1:-1:color=black,fps=25",
                "-c:v", "libx264", "-pix_fmt", "yuv420p", "-c:a", "aac", "-ar", "44100",
                "-preset", "fast", "-crf", "23", str(out)], capture_output=True)
            if out.exists() and out.stat().st_size > 0:
                dur = e - s
                sliced.append({"file": out.name, "title": c.get("title", ""),
                               "start": cursor, "end": cursor + dur})
                cuts.append({"start": cursor, "end": cursor + dur,
                             "title": c.get("title", "")})
                cursor += dur
        if not sliced:
            tg(f"❌ {proj['name']}: no usable cuts."); log(jid, 0, "no cuts", status="error"); return
        wdir = WORK / jid
        (wdir / "concat.txt").write_text("".join(f"file '{EXPORT / s['file']}'\n" for s in sliced))
        final_vid = EXPORT / f"{jid}_{slug}.mp4"
        subprocess.run(["ffmpeg", "-y", "-f", "concat", "-safe", "0",
                        "-i", str(wdir / "concat.txt"),
                        "-vf", "scale=1920:1080:force_original_aspect_ratio=decrease,pad=1920:1080:(ow-iw)/2:(oh-ih)/2",
                        "-c:v", "libx264", "-profile:v", "high", "-level", "4.0",
                        "-pix_fmt", "yuv420p", "-c:a", "aac", "-b:a", "128k",
                        "-preset", "fast", "-crf", "23", str(final_vid)],
                       capture_output=True)

        # Voice per voice_mode: mine = original audio only, ai/both = TTS
        tts_file = None
        if voice_mode in ("ai", "both"):
            log(jid, 84, "Generating AI voice…")
            voice_path = proj.get("voice_path") or _channel_voice(channels)
            try:
                wav = step_tts(script, voice_path)
                tts_file = str(EXPORT / f"{jid}_script.wav")
                Path(tts_file).write_bytes(wav)
            except Exception as e:
                log(jid, 85, f"TTS failed ({e})")

        log(jid, 90, "Channel export…")
        ch_map = {"mine": ["A"], "ai": ["B"], "both": ["A", "B", "C"]}
        channel_job = step_channel_export(str(final_vid), tts_file,
                                          clone_text=script,
                                          export_channels=ch_map.get(voice_mode, ["C"]))

        # Captions across new timeline from a-roll segments mapped by cut source
        segments_out = []
        sentences = [x.strip() for x in re.split(r"(?<=[.!?])\s+", script) if x.strip()]
        if sentences and cursor > 0:
            per = cursor / len(sentences)
            segments_out = [{"start": i * per, "end": (i + 1) * per, "text": s}
                            for i, s in enumerate(sentences)]

        brain_pkg = {"topic": proj["name"], "summary": brain.get("title", ""),
                     "hook": brain.get("hook", ""), "script": script,
                     "tags": brain.get("tags", []), "best_clips": cuts}
        gdrive_path = build_asset_package(
            jid, str(final_vid), brain_pkg,
            " ".join(c["transcript"] for c in clip_info), segments_out,
            [s for c in clip_info for s in c["scenes"]],
            sliced, tts_file, ch_names, sam_notes=extra,
            priority=proj.get("priority", "normal"),
            deadline=proj.get("deadline", ""), raw_gdrive=raw_gdrive)

        draft = {"job_id": jid, "file": proj.get("video_name") or proj["name"],
                 "video_path": str(final_vid), "created": time.strftime("%Y-%m-%d %H:%M"),
                 "duration": cursor, "channel": ch_names,
                 "summary": brain.get("title", ""), "topic": proj["name"],
                 "hook": brain.get("hook", ""), "script": script,
                 "tags": brain.get("tags", []), "cuts": cuts,
                 "sam_notes": extra, "transcript": script,
                 "segments": segments_out, "scenes": [],
                 "sliced_clips": sliced, "tts_file": tts_file, "status": "ready"}
        save_json(DRAFTS_DIR / f"{jid}.json", draft)

        excel_log({"Date": time.strftime("%Y-%m-%d"), "Time": time.strftime("%H:%M"),
                   "Job ID": jid, "File": f"[PROJECT] {proj['name']}",
                   "Channel": ch_names, "Duration(s)": round(cursor, 1),
                   "Transcript Preview": script[:150], "Scenes": len(clip_info),
                   "Clips": len(sliced), "TTS": "yes" if tts_file else "no",
                   "Export": "yes" if channel_job and "error" not in channel_job else "no",
                   "Status": "done", "Sam Notes": extra[:200],
                   "Google Drive Path": gdrive_path})

        proj["status"] = "done"; proj["last_job"] = jid
        projects[pid] = proj; save_json(PROJECTS_FILE, projects)

        tg(f"✅ <b>{proj['name']} ready!</b>\n"
           f"\U0001f4a1 {brain.get('title','')[:60]}\n"
           f"✂️ {len(sliced)} cuts from {len(clip_info)} clips | voice: {voice_mode}\n\n"
           f"\U0001f39e <b>Open on OpenCut timeline:</b>\n"
           f"http://192.168.0.78:9500/ai-draft/{jid}")
        log(jid, 100, "✓ Project video ready!", status="done",
            result={"title": brain.get("title", ""), "video": final_vid.name,
                    "clips": len(sliced)})
    except Exception as e:
        import traceback
        tg(f"❌ <b>Project error</b>\n{e}")
        log(jid, 0, f"Error: {e}", status="error", trace=traceback.format_exc())

# ── Recall & Remix engine ─────────────────────────────────────────────────────
def brain_remix(request_text, assets):
    """Pick the best moments across old assets and write a new script."""
    asset_blocks = []
    for rec in assets:
        segs = rec.get("segments") or []
        seg_lines = "\n".join(
            f"    [{s.get('start',0):.1f}s–{s.get('end',0):.1f}s] {s.get('text','').strip()}"
            for s in segs[:40])
        asset_blocks.append(
            f"ASSET {rec['jid']} — {rec.get('topic','')} ({rec.get('file','')}):\n"
            f"  summary: {rec.get('summary','')[:200]}\n"
            f"  spoken segments:\n{seg_lines or '    (no speech)'}")
    prompt = f"""Creator request: {request_text}

You have these OLD videos available (with word-accurate timestamps):

{chr(10).join(asset_blocks)}

Build a NEW short video by picking the best moments. Return ONLY valid JSON:
{{
  "title": "new video title",
  "hook": "grabby opening line",
  "script": "45-60s voiceover script tying the old clips together, natural for TTS",
  "clips": [
    {{"asset": "<asset jid>", "start": 0.0, "end": 8.0, "title": "why this moment"}}
  ],
  "tags": ["tag1","tag2","tag3"]
}}
Rules: 2-6 clips, each 3-20 seconds, only use timestamps that exist in the segments above."""
    last_err = None
    for attempt in range(3):
        raw = groq_chat([
            {"role": "system", "content": "You are a video remix editor AI. "
             "Respond with ONLY a JSON object — no markdown, no explanation."},
            {"role": "user", "content": prompt}
        ])
        raw = re.sub(r"^```json?\s*", "", raw.strip())
        raw = re.sub(r"\s*```$", "", raw.strip())
        try:
            return json.loads(raw)
        except json.JSONDecodeError as e:
            last_err = e
        # Fallback: extract the outermost JSON object from surrounding prose
        m = re.search(r"\{.*\}", raw, re.S)
        if m:
            try:
                return json.loads(m.group(0))
            except json.JSONDecodeError as e:
                last_err = e
    raise ValueError(f"Brain returned unparseable JSON after 3 tries: {last_err}")

def run_remix(jid, request_text):
    try:
        wdir = WORK / jid; wdir.mkdir(exist_ok=True)

        log(jid, 5, "Searching Hermes memory index…")
        assets = search_assets(request_text, top=3)
        if not assets:
            tg(f"\U0001f50d <b>Remix:</b> nothing in my memory matches "
               f"“{request_text[:80]}”. Try different words or /find first.")
            log(jid, 0, "No matching assets", status="error")
            return
        names = ", ".join(a.get("topic") or a.get("file", "?") for a in assets)
        tg(f"\U0001f3ac <b>Remix started</b>\nFound {len(assets)} old video(s): {names}\n"
           f"Pulling footage + building new edit…")

        log(jid, 15, f"Recalling {len(assets)} asset(s) (Drive if archived)…")
        sources = {}
        for rec in assets:
            p = ensure_asset_local(rec)
            if p:
                sources[rec["jid"]] = p
        if not sources:
            tg("❌ <b>Remix:</b> found matches but couldn't get the video files "
               "(Google Drive not connected yet and local copies are gone).")
            log(jid, 0, "No sources available", status="error")
            return
        assets = [a for a in assets if a["jid"] in sources]

        log(jid, 35, "Groq Brain: picking best moments + writing script…")
        brain = brain_remix(request_text, assets)
        script = brain.get("script", "")

        log(jid, 55, "Slicing clips from old footage…")
        sliced, cuts, cursor = [], [], 0.0
        for i, c in enumerate(brain.get("clips", [])[:6]):
            src = sources.get(str(c.get("asset", "")).strip())
            if not src:
                continue
            s, e = float(c.get("start", 0)), float(c.get("end", 0))
            if e - s < 1:
                continue
            out = EXPORT / f"{jid}_clip{i+1}.mp4"
            subprocess.run([
                "ffmpeg", "-y", "-i", str(src), "-ss", str(s), "-to", str(e),
                "-vf", "scale=1280:720:force_original_aspect_ratio=decrease,"
                       "pad=1280:720:-1:-1:color=black,fps=25",
                "-c:v", "libx264", "-pix_fmt", "yuv420p", "-c:a", "aac", "-ar", "44100",
                "-preset", "fast", "-crf", "23", str(out)
            ], capture_output=True)
            if out.exists() and out.stat().st_size > 0:
                dur = e - s
                sliced.append({"file": out.name, "title": c.get("title", f"Clip {i+1}"),
                               "start": cursor, "end": cursor + dur})
                cuts.append({"start": cursor, "end": cursor + dur,
                             "title": c.get("title", "")})
                cursor += dur
        if not sliced:
            tg("❌ <b>Remix:</b> couldn't slice any usable clips.")
            log(jid, 0, "No clips sliced", status="error")
            return

        log(jid, 70, "Generating voiceover (XTTS)…")
        tts_file = None
        try:
            wav = step_tts(script or request_text)
            tts_file = str(EXPORT / f"{jid}_script.wav")
            Path(tts_file).write_bytes(wav)
        except Exception as e:
            log(jid, 72, f"TTS failed ({e})")

        log(jid, 80, "Assembling remix video…")
        concat_list = wdir / "concat.txt"
        concat_list.write_text("".join(f"file '{EXPORT / c['file']}'\n" for c in sliced))
        final_vid = EXPORT / f"{jid}_remix.mp4"
        subprocess.run([
            "ffmpeg", "-y", "-f", "concat", "-safe", "0", "-i", str(concat_list),
            "-vf", "scale=1920:1080:force_original_aspect_ratio=decrease,pad=1920:1080:(ow-iw)/2:(oh-ih)/2",
            "-c:v", "libx264", "-profile:v", "high", "-level", "4.0",
            "-pix_fmt", "yuv420p", "-c:a", "aac", "-b:a", "128k",
            "-preset", "fast", "-crf", "23", str(final_vid)
        ], capture_output=True)

        log(jid, 88, "Channel export A/B/C…")
        channel_job = None
        if tts_file and final_vid.exists():
            channel_job = step_channel_export(str(final_vid), tts_file,
                                              clone_text=script)

        # Captions: spread the new script across the remix timeline
        segments = []
        sentences = [x.strip() for x in re.split(r"(?<=[.!?])\s+", script) if x.strip()]
        if sentences and cursor > 0:
            per = cursor / len(sentences)
            segments = [{"start": i * per, "end": (i + 1) * per, "text": s}
                        for i, s in enumerate(sentences)]

        draft = {
            "job_id": jid, "file": f"[REMIX] {brain.get('title', request_text[:40])}",
            "video_path": str(final_vid), "created": time.strftime("%Y-%m-%d %H:%M"),
            "duration": cursor, "channel": "remix",
            "summary": brain.get("title", ""), "topic": brain.get("title", ""),
            "hook": brain.get("hook", ""), "script": script,
            "tags": brain.get("tags", []), "cuts": cuts, "sam_notes": request_text,
            "transcript": script, "segments": segments,
            "scenes": [], "sliced_clips": sliced,
            "tts_file": tts_file, "status": "ready",
        }
        save_json(DRAFTS_DIR / f"{jid}.json", draft)

        excel_log({
            "Date": time.strftime("%Y-%m-%d"), "Time": time.strftime("%H:%M"),
            "Job ID": jid, "File": f"[REMIX] {brain.get('title','')[:60]}",
            "Channel": "remix", "Duration(s)": round(cursor, 1),
            "Transcript Preview": script[:150],
            "Scenes": 0, "Clips": len(sliced),
            "TTS": "yes" if tts_file else "no",
            "Export": "yes" if channel_job and "error" not in channel_job else "no",
            "Status": "done", "Sam Notes": request_text[:200],
            "Google Drive Path": "",
        })

        result = {
            "title": brain.get("title", ""), "hook": brain.get("hook", ""),
            "script": script, "tags": brain.get("tags", []),
            "best_clips": sliced, "tts_file": tts_file,
            "video": final_vid.name, "channel_export": channel_job,
            "source_assets": [a["jid"] for a in assets],
        }
        tg(
            f"✅ <b>Remix ready!</b>\n"
            f"\U0001f4a1 {brain.get('title','')[:60]}\n"
            f"✂️ {len(sliced)} clips from {len(assets)} old video(s)\n"
            f"\U0001f3b5 {brain.get('hook','')[:80]}\n\n"
            f"\U0001f39e <b>Open on OpenCut timeline:</b>\n"
            f"http://192.168.0.78:9500/ai-draft/{jid}"
        )
        log(jid, 100, "✓ Remix ready!", status="done", result=result)

    except Exception as e:
        import traceback
        tg(f"❌ <b>Remix error</b>\n{e}")
        log(jid, 0, f"Error: {e}", status="error", trace=traceback.format_exc())

def start_remix(request_text):
    jid = uuid.uuid4().hex[:10]
    jobs[jid] = {"status": "running", "progress": 0, "msg": "Remix queued…"}
    threading.Thread(target=run_remix, args=(jid, request_text), daemon=True).start()
    return jid

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory(".", "auto.html")

@app.route("/process", methods=["POST"])
def process():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Empty filename"}), 400
    jid = uuid.uuid4().hex[:10]
    vpath = UPLOAD / f"{jid}_{secure_filename(f.filename)}"
    f.save(str(vpath))
    options = {
        "voice_path":          request.form.get("voice_path", ""),
        "channel_ids":         json.loads(request.form.get("channel_ids", "[]")),
        "save_blueprint_name": request.form.get("save_blueprint_name", ""),
        "priority":            request.form.get("priority", "normal"),
        "deadline":            request.form.get("deadline", ""),
    }
    bp_id = request.form.get("blueprint_id", "")
    if bp_id:
        bps = load_json(BLUEPRINTS_FILE)
        if bp_id in bps:
            bp = bps[bp_id]
            options.setdefault("voice_path", bp.get("voice_path", ""))
            if not options["channel_ids"]:
                options["channel_ids"] = bp.get("channel_ids", [])
    jobs[jid] = {"status": "running", "progress": 0, "msg": "Queued…"}
    threading.Thread(target=run_pipeline, args=(jid, vpath, options), daemon=True).start()
    return jsonify({"job_id": jid})

@app.route("/auto-shorts", methods=["POST"])
def auto_shorts():
    data = request.get_json(force=True)
    idea = (data.get("idea") or "").strip()
    if not idea:
        return jsonify({"error": "No idea provided"}), 400
    jid = uuid.uuid4().hex[:10]
    options = {
        "voice_path":          data.get("voice_path", ""),
        "channel_ids":         data.get("channel_ids", []),
        "save_blueprint_name": data.get("save_blueprint_name", ""),
    }
    bp_id = data.get("blueprint_id", "")
    if bp_id:
        bps = load_json(BLUEPRINTS_FILE)
        if bp_id in bps:
            bp = bps[bp_id]
            options.setdefault("voice_path", bp.get("voice_path", ""))
            if not options["channel_ids"]:
                options["channel_ids"] = bp.get("channel_ids", [])
    jobs[jid] = {"status": "running", "progress": 0, "msg": "Starting…"}
    threading.Thread(target=run_auto_shorts, args=(jid, idea, options), daemon=True).start()
    return jsonify({"job_id": jid})

@app.route("/job/<jid>")
def job_status(jid):
    j = jobs.get(jid)
    if not j:
        return jsonify({"error": "Not found"}), 404
    return jsonify(j)

# ── Projects API ──────────────────────────────────────────────────────────────
@app.route("/projects", methods=["GET"])
def list_projects():
    projects = load_json(PROJECTS_FILE)
    for pid, p in projects.items():
        p["counts"] = project_counts(p)
    return jsonify(projects)

@app.route("/projects", methods=["POST"])
def create_project():
    data = request.get_json(force=True)
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Project name required"}), 400
    projects = load_json(PROJECTS_FILE)
    pid = uuid.uuid4().hex[:8]
    vslug = re.sub(r"[^a-z0-9]+", "_",
                   (data.get("video_name") or name).lower())[:40].strip("_") or pid
    slug = vslug
    if any(p["slug"] == slug for p in projects.values()):
        slug = f"{slug}_{pid[:4]}"
    # Live inside the channel inbox when a channel is chosen (Sam's structure);
    # the watcher skips project folders via project_brief.json marker.
    channel_ids = data.get("channel_ids", [])
    chs = load_json(CHANNELS_FILE)
    ch_folder = ""
    if channel_ids and channel_ids[0] in chs:
        ch_folder = chs[channel_ids[0]].get("folder_name") or \
            chs[channel_ids[0]]["name"].lower().replace(" ", "_").replace("/", "-")
    if ch_folder:
        pdir = CHANNELS_ROOT / ch_folder / "inbox" / slug
        smb  = f"\\\\192.168.0.78\\studio\\channels\\{ch_folder}\\inbox\\{slug}"
        smb_url = f"smb://192.168.0.78/studio/channels/{ch_folder}/inbox/{slug}"
    else:
        pdir = PROJECTS_ROOT / slug
        smb  = f"\\\\192.168.0.78\\studio\\projects\\{slug}"
        smb_url = f"smb://192.168.0.78/studio/projects/{slug}"
    for sub in PROJECT_SUBS:
        (pdir / sub).mkdir(parents=True, exist_ok=True)
    projects[pid] = {
        "id": pid, "name": name, "slug": slug, "dir": str(pdir),
        "smb": smb, "smb_url": smb_url,
        "bio": data.get("bio", ""),
        "video_name":  data.get("video_name", ""),
        "channel_ids": data.get("channel_ids", []),
        "blueprint_id": data.get("blueprint_id", ""),
        "voice_mode":  data.get("voice_mode", "both"),   # ai | mine | both
        "voice_path":  data.get("voice_path", ""),
        "deadline":    data.get("deadline", ""),
        "priority":    data.get("priority", "normal"),
        "notes":       data.get("notes", ""),
        "status":      "awaiting footage",
        "created":     time.strftime("%Y-%m-%d %H:%M"),
    }
    # Blueprint fills gaps
    bp_id = data.get("blueprint_id", "")
    if bp_id:
        bps = load_json(BLUEPRINTS_FILE)
        if bp_id in bps:
            bp = bps[bp_id]
            if not projects[pid]["voice_path"]:
                projects[pid]["voice_path"] = bp.get("voice_path", "")
            if not projects[pid]["channel_ids"]:
                projects[pid]["channel_ids"] = bp.get("channel_ids", [])
    save_json(PROJECTS_FILE, projects)
    # Brief lives inside the folder too (also tells the watcher: hands off)
    (pdir / "project_brief.json").write_text(json.dumps(projects[pid], indent=2))
    tg(f"\U0001f4c1 <b>New project: {name}</b>\n"
       f"Drop your files in:\n<code>{smb}</code>\n"
       f"(a-roll / b-roll / pics / other)\nThen hit AI Check on the site.")
    return jsonify({"ok": True, "id": pid, "slug": slug,
                    "smb": smb, "smb_url": smb_url})

@app.route("/projects/<pid>/upload", methods=["POST"])
def project_upload(pid):
    projects = load_json(PROJECTS_FILE)
    proj = projects.get(pid)
    if not proj:
        return jsonify({"error": "Not found"}), 404
    sub = request.form.get("folder", "other")
    if sub not in PROJECT_SUBS:
        sub = "other"
    saved = 0
    for f in request.files.getlist("file"):
        if f.filename:
            f.save(str(project_dir(proj) / sub / secure_filename(f.filename)))
            saved += 1
    return jsonify({"ok": True, "saved": saved,
                    "counts": project_counts(proj)})

@app.route("/projects/<pid>/check", methods=["POST"])
def project_check(pid):
    projects = load_json(PROJECTS_FILE)
    proj = projects.get(pid)
    if not proj:
        return jsonify({"error": "Not found"}), 404
    jid = uuid.uuid4().hex[:10]
    proj["status"] = "processing"; proj["last_job"] = jid
    save_json(PROJECTS_FILE, projects)
    jobs[jid] = {"status": "running", "progress": 0, "msg": "AI Check queued…"}
    threading.Thread(target=run_project_check, args=(jid, pid), daemon=True).start()
    return jsonify({"ok": True, "job_id": jid})

@app.route("/projects/<pid>", methods=["DELETE"])
def delete_project(pid):
    projects = load_json(PROJECTS_FILE)
    proj = projects.pop(pid, None)
    save_json(PROJECTS_FILE, projects)
    if proj:
        shutil.rmtree(project_dir(proj), ignore_errors=True)
    return jsonify({"ok": True})

# ── Blueprints ────────────────────────────────────────────────────────────────
@app.route("/blueprints", methods=["GET"])
def list_blueprints():
    return jsonify(load_json(BLUEPRINTS_FILE))

@app.route("/blueprints", methods=["POST"])
def save_blueprint():
    data = request.get_json(force=True)
    if not data.get("name"):
        return jsonify({"error": "Name required"}), 400
    bps = load_json(BLUEPRINTS_FILE)
    bid = data.get("id") or uuid.uuid4().hex[:8]
    bps[bid] = {
        "id": bid, "name": data["name"],
        "voice_path":  data.get("voice_path", ""),
        "channel_ids": data.get("channel_ids", []),
        "topic":       data.get("topic", ""),
        "notes":       data.get("notes", ""),
        "created":     data.get("created", time.strftime("%Y-%m-%d %H:%M")),
        "updated":     time.strftime("%Y-%m-%d %H:%M"),
    }
    save_json(BLUEPRINTS_FILE, bps)
    return jsonify({"ok": True, "id": bid})

@app.route("/blueprints/<bid>", methods=["DELETE"])
def delete_blueprint(bid):
    bps = load_json(BLUEPRINTS_FILE)
    bps.pop(bid, None)
    save_json(BLUEPRINTS_FILE, bps)
    return jsonify({"ok": True})

# ── Channels ──────────────────────────────────────────────────────────────────
@app.route("/channels", methods=["GET"])
def list_channels():
    return jsonify(load_json(CHANNELS_FILE))

@app.route("/channels", methods=["POST"])
def save_channel():
    data = request.get_json(force=True)
    if not data.get("name"):
        return jsonify({"error": "Name required"}), 400
    chs = load_json(CHANNELS_FILE)
    if len(chs) >= 50 and not data.get("id"):
        return jsonify({"error": "Maximum 50 channels"}), 400
    cid = data.get("id") or uuid.uuid4().hex[:8]
    safe = data["name"].lower().replace(" ", "_").replace("/", "-")
    chs[cid] = {
        "id": cid, "name": data["name"], "folder_name": safe,
        "platform":            data.get("platform", "tiktok"),
        "url":                 data.get("url", ""),
        "voice_path":          data.get("voice_path", ""),
        "composio_connection": data.get("composio_connection", ""),
        "style":               data.get("style", ""),
        "format":              data.get("format", "both"),
        "auto_mode":           bool(data.get("auto_mode", False)),
        "agent_role":          data.get("agent_role", "mini"),
        "audio_variant":       data.get("audio_variant", "C"),
        "linked_channel":      data.get("linked_channel", ""),
        "linked_notes":        data.get("linked_notes", ""),
        "language":            data.get("language", "en"),
        "stagger_hours":       float(data.get("stagger_hours", 0) or 0),
        "niche":               data.get("niche", ""),
        "notes":               data.get("notes", ""),
        "created":             data.get("created", time.strftime("%Y-%m-%d %H:%M")),
    }
    save_json(CHANNELS_FILE, chs)
    for sub in ("inbox", "exports"):
        (CHANNELS_ROOT / safe / sub).mkdir(parents=True, exist_ok=True)
    try:
        import urllib.request as _ur
        _ur.urlopen("http://localhost:9531/refresh-folders", timeout=2)
    except Exception:
        pass
    return jsonify({"ok": True, "id": cid,
                    "smb_path": f"\\\\192.168.0.78\\studio\\channels\\{safe}\\inbox"})

@app.route("/channels/<cid>", methods=["DELETE"])
def delete_channel(cid):
    chs = load_json(CHANNELS_FILE)
    chs.pop(cid, None)
    save_json(CHANNELS_FILE, chs)
    return jsonify({"ok": True})

# ── Shared ────────────────────────────────────────────────────────────────────
@app.route("/voices")
def voices():
    import urllib.request
    try:
        with urllib.request.urlopen(f"{CHANNEL_URL}/voices", timeout=5) as r:
            return r.read(), 200, {"Content-Type": "application/json"}
    except Exception:
        return jsonify({"voices": {}})

@app.route("/exports/<path:fn>")
def serve_export(fn):
    return send_from_directory(str(EXPORT), fn)

@app.route("/drafts")
def list_drafts():
    out = []
    for f in sorted(DRAFTS_DIR.glob("*.json"), reverse=True)[:50]:
        try:
            d = json.loads(f.read_text())
            out.append(d)
        except Exception:
            pass
    return jsonify({"drafts": out})

@app.route("/drafts/<jid>")
def get_draft(jid):
    f = DRAFTS_DIR / f"{jid}.json"
    if not f.exists():
        return jsonify({"error": "Not found"}), 404
    return jsonify(json.loads(f.read_text()))

@app.route("/drafts/<jid>", methods=["DELETE"])
def delete_draft(jid):
    (DRAFTS_DIR / f"{jid}.json").unlink(missing_ok=True)
    return jsonify({"ok": True})

@app.route("/excel")
def download_excel():
    if EXCEL_LOG.exists():
        return send_from_directory(str(DATA_DIR), "pipeline_log.xlsx",
                                   as_attachment=True)
    return jsonify({"error": "No log yet"}), 404

# ── Settings (master Composio for planning etc.) ─────────────────────────────
@app.route("/settings", methods=["GET"])
def get_settings():
    return jsonify(load_json(SETTINGS_FILE))

@app.route("/settings", methods=["POST"])
def save_settings():
    s = load_json(SETTINGS_FILE)
    s.update(request.get_json(force=True))
    save_json(SETTINGS_FILE, s)
    return jsonify({"ok": True})

# ── Hermes control webhook ────────────────────────────────────────────────────
@app.route("/control", methods=["POST"])
def hermes_control():
    data = request.get_json(force=True)
    action = data.get("action", "").lower()
    params = data.get("params", {})

    if action == "auto_shorts":
        idea = params.get("idea", "")
        if not idea:
            return jsonify({"error": "idea required"}), 400
        jid = uuid.uuid4().hex[:10]
        options = {"voice_path": params.get("voice_path", ""),
                   "channel_ids": params.get("channel_ids", [])}
        jobs[jid] = {"status": "running", "progress": 0, "msg": "Queued by Hermes…"}
        threading.Thread(target=run_auto_shorts, args=(jid, idea, options), daemon=True).start()
        return jsonify({"ok": True, "job_id": jid})

    elif action == "list_jobs":
        recent = [{k: v[k] for k in ("status", "progress", "msg") if k in v} | {"jid": k}
                  for k, v in list(jobs.items())[-10:]]
        return jsonify({"jobs": recent})

    elif action == "list_channels":
        return jsonify(load_json(CHANNELS_FILE))

    elif action == "list_blueprints":
        return jsonify(load_json(BLUEPRINTS_FILE))

    elif action == "status":
        return jsonify({"ok": True,
                        "active_jobs": sum(1 for j in jobs.values() if j.get("status") == "running")})

    elif action == "telegram":
        msg = params.get("message", "")
        if msg:
            tg(msg)
        return jsonify({"ok": True})

    elif action == "search":
        return jsonify({"results": search_assets(params.get("query", ""), top=5)})

    elif action == "remix":
        req_text = params.get("request", "") or params.get("query", "")
        if not req_text:
            return jsonify({"error": "request required"}), 400
        return jsonify({"ok": True, "job_id": start_remix(req_text)})

    else:
        return jsonify({"error": f"Unknown action: {action}",
                        "actions": ["auto_shorts", "list_jobs", "list_channels",
                                    "list_blueprints", "status", "telegram",
                                    "search", "remix"]}), 400

# ── Clean + Assemble (Gling/Magisto style) ─────────────────────────────────

FILLER_WORDS_STRICT = {"um","uh","er","ah","hmm"}
FILLER_WORDS_MEDIUM = {*FILLER_WORDS_STRICT,
    "like","literally","basically","you know","i mean",
    "sort of","kind of","actually","honestly","obviously"}

clean_jobs = {}

def detect_duplicates(clips):
    import difflib
    groups = []
    used = set()
    for i, ci in enumerate(clips):
        if i in used:
            continue
        group = [i]
        ti = ci.get("transcript","")[:250].lower()
        for j, cj in enumerate(clips):
            if j <= i or j in used:
                continue
            tj = cj.get("transcript","")[:250].lower()
            ratio = difflib.SequenceMatcher(None, ti, tj).ratio()
            if ratio > 0.55:
                group.append(j)
                used.add(j)
        if len(group) > 1:
            used.add(i)
            groups.append(group)
    return groups

def build_edit_plan(words, silence_thresh=0.40, filler_mode="medium"):
    fillers = FILLER_WORDS_MEDIUM if filler_mode=="medium" else \
              FILLER_WORDS_STRICT if filler_mode=="strict" else set()
    segments = []
    n = len(words)
    i = 0
    last_end = 0.0
    while i < n:
        w = words[i]
        raw = w["word"].strip()
        norm = raw.lower().strip(".,!?;:-\"'")
        gap = w["start"] - last_end
        if i > 0 and gap >= silence_thresh:
            segments.append({"id":len(segments),"start":round(last_end,3),
                             "end":round(w["start"],3),
                             "text":f"[{gap:.1f}s pause]",
                             "keep":False,"reason":"silence","gap":round(gap,3)})
        if norm in fillers:
            segments.append({"id":len(segments),"start":round(w["start"],3),
                             "end":round(w["end"],3),"text":raw,
                             "keep":False,"reason":"filler"})
            last_end = w["end"]; i += 1; continue
        seg_words=[raw]; seg_start=w["start"]; seg_end=w["end"]
        j = i+1
        while j < n:
            nw=words[j]; nr=nw["word"].strip()
            nn=nr.lower().strip(".,!?;:-\"'")
            if (nw["start"]-seg_end)>=silence_thresh or nn in fillers:
                break
            seg_words.append(nr); seg_end=nw["end"]; j+=1
        segments.append({"id":len(segments),"start":round(seg_start,3),
                         "end":round(seg_end,3),"text":" ".join(seg_words),
                         "keep":True,"reason":"speech"})
        last_end=seg_end; i=j
    return segments

def _transcribe_video(video_path):
    import urllib.request as _ur, base64 as _b64
    ap = Path(str(video_path)+".wav")
    subprocess.run(["ffmpeg","-y","-i",str(video_path),"-ar","16000","-ac","1",str(ap)],
                   capture_output=True)
    boundary = uuid.uuid4().hex
    with open(str(ap),"rb") as fh: audio=fh.read()
    body=(f"--{boundary}\r\nContent-Disposition: form-data; name=\"file\"; "
          f"filename=\"audio.wav\"\r\nContent-Type: audio/wav\r\n\r\n"
          ).encode()+audio+f"\r\n--{boundary}--\r\n".encode()
    req=_ur.Request(f"{WHISPER_URL}/transcribe",data=body,
                    headers={"Content-Type":f"multipart/form-data; boundary={boundary}"})
    with _ur.urlopen(req,timeout=600) as r: stt=json.load(r)
    ap.unlink(missing_ok=True)
    words=[]
    for seg in stt.get("segments",[]):
        for w in seg.get("words",[]):
            if w.get("word","").strip():
                words.append({"word":w["word"],"start":float(w["start"]),"end":float(w["end"])})
    if not words:
        for seg in stt.get("segments",[]):
            words.append({"word":seg.get("text","").strip(),
                          "start":float(seg["start"]),"end":float(seg["end"])})
    return words, stt.get("text",""), float(stt.get("duration",0) or 0)

@app.route("/clean/upload", methods=["POST"])
def clean_upload():
    files=request.files.getlist("files") or \
          ([request.files["file"]] if "file" in request.files else [])
    if not files: return jsonify({"error":"No files"}),400
    silence_thresh=float(request.form.get("silence_thresh",0.40))
    filler_mode=request.form.get("filler_mode","medium")
    jid=uuid.uuid4().hex[:10]
    work=Path(f"/opt/studio/media/clean/{jid}"); work.mkdir(parents=True,exist_ok=True)
    clips_meta=[]
    for i,f in enumerate(files):
        ext=Path(f.filename).suffix or ".mp4"
        vp=work/f"clip_{i:03d}{ext}"; f.save(str(vp))
        clips_meta.append({"idx":i,"filename":f.filename,"path":str(vp),
                           "status":"queued","transcript":"","words":[],
                           "duration":0,"thumb":"","segments":[]})
    clean_jobs[jid]={"status":"processing","progress":0,"clips":clips_meta,
                     "silence_thresh":silence_thresh,"filler_mode":filler_mode,
                     "duplicates":[],"order":list(range(len(clips_meta)))}
    def _process():
        total=len(clips_meta)
        for ci,clip in enumerate(clips_meta):
            try:
                thumb_path=work/f"thumb_{ci}.jpg"
                subprocess.run(["ffmpeg","-y","-i",clip["path"],"-ss","00:00:01",
                                "-vframes","1","-vf","scale=280:-1",str(thumb_path)],
                               capture_output=True)
                if thumb_path.exists():
                    import base64 as b64
                    clip["thumb"]="data:image/jpeg;base64,"+b64.b64encode(
                        thumb_path.read_bytes()).decode()
                words,text,dur=_transcribe_video(clip["path"])
                clip.update({"words":words,"transcript":text,"duration":dur,
                              "status":"transcribed"})
                clip["segments"]=build_edit_plan(words,
                    clean_jobs[jid]["silence_thresh"],clean_jobs[jid]["filler_mode"])
                clean_jobs[jid]["progress"]=int((ci+1)/total*85)
            except Exception as e:
                clip["status"]="error"; clip["error"]=str(e)
        clean_jobs[jid]["duplicates"]=detect_duplicates(clips_meta)
        clean_jobs[jid]["status"]="ready"; clean_jobs[jid]["progress"]=100
    threading.Thread(target=_process,daemon=True).start()
    return jsonify({"jid":jid})

@app.route("/clean/<jid>")
def clean_status(jid):
    job=clean_jobs.get(jid)
    if not job: return jsonify({"error":"not found"}),404
    out={k:v for k,v in job.items() if k!="clips"}
    out["clips"]=[{k:v for k,v in c.items() if k!="words"} for c in job.get("clips",[])]
    return jsonify(out)

@app.route("/clean/<jid>/reanalyse",methods=["POST"])
def clean_reanalyse(jid):
    job=clean_jobs.get(jid)
    if not job: return jsonify({"error":"not found"}),404
    data=request.get_json(force=True)
    thresh=float(data.get("silence_thresh",0.40))
    fmode=data.get("filler_mode",job.get("filler_mode","medium"))
    clean_jobs[jid]["silence_thresh"]=thresh; clean_jobs[jid]["filler_mode"]=fmode
    for clip in job.get("clips",[]):
        clip["segments"]=build_edit_plan(clip.get("words",[]),thresh,fmode)
    return jsonify({"clips":[{k:v for k,v in c.items() if k!="words"}
                              for c in job.get("clips",[])]})

@app.route("/clean/<jid>/approve",methods=["POST"])
def clean_approve(jid):
    job=clean_jobs.get(jid)
    if not job: return jsonify({"error":"not found"}),404
    data=request.get_json(force=True)
    clean_jobs[jid]["order"]=data.get("order",job["order"])
    for upd in data.get("segments",[]):
        idx=upd.get("clip_idx")
        if idx is not None and 0<=idx<len(job["clips"]):
            job["clips"][idx]["segments"]=upd["segments"]
    clean_jobs[jid]["approved"]=True
    return jsonify({"ok":True})

@app.route("/clean/<jid>/render",methods=["POST"])
def clean_render(jid):
    job=clean_jobs.get(jid)
    if not job: return jsonify({"error":"not found"}),404
    clean_jobs[jid]["status"]="rendering"; clean_jobs[jid]["render_progress"]=0
    def _render():
        try:
            work=Path(f"/opt/studio/media/clean/{jid}")
            order=clean_jobs[jid].get("order",list(range(len(job["clips"]))))
            all_files=[]
            for oi,ci in enumerate(order):
                clip=job["clips"][ci]; vp=Path(clip["path"])
                keeps=[s for s in clip.get("segments",[]) if s.get("keep",True)]
                if not keeps: keeps=[{"start":0,"end":clip.get("duration",0)}]
                for ki,s in enumerate(keeps):
                    cp=work/f"out_{oi:03d}_{ki:04d}.mp4"
                    subprocess.run(["ffmpeg","-y","-ss",str(s["start"]),
                                    "-to",str(s["end"]),"-i",str(vp),
                                    "-c","copy",str(cp)],capture_output=True)
                    if cp.exists() and cp.stat().st_size>500: all_files.append(cp)
            clean_jobs[jid]["render_progress"]=60
            concat=work/"concat.txt"
            concat.write_text("\n".join(f"file '{c}'" for c in all_files))
            out=EXPORT/f"clean_{jid}.mp4"
            subprocess.run(["ffmpeg","-y","-f","concat","-safe","0",
                             "-i",str(concat),
                             "-vf","scale=1920:1080:force_original_aspect_ratio=decrease,pad=1920:1080:(ow-iw)/2:(oh-ih)/2",
                             "-c:v","libx264","-profile:v","high","-level","4.0",
                             "-pix_fmt","yuv420p","-c:a","aac","-b:a","128k",
                             "-preset","fast","-crf","23",str(out)],capture_output=True)
            clean_jobs[jid].update({"status":"done","render_progress":100,
                                     "output":f"/exports/clean_{jid}.mp4",
                                     "output_name":f"clean_{jid}.mp4"})
        except Exception as e:
            clean_jobs[jid].update({"status":"error","error":str(e)})
    threading.Thread(target=_render,daemon=True).start()
    return jsonify({"ok":True})

@app.route("/tg-wait-reply", methods=["POST"])
def tg_wait_reply():
    # Blocking endpoint: send a TG prompt, wait up to timeout secs for Sam's reply.
    # Called by the watcher so it doesn't need its own getUpdates loop (avoids 409).
    data = request.get_json(force=True) or {}
    prompt  = data.get("prompt", "")
    timeout = int(data.get("timeout", 300))

    if prompt:
        tg(prompt)

    import uuid as _uuid
    key = "watcher-" + _uuid.uuid4().hex[:8]
    import threading as _th
    ev = _th.Event()
    with _tg_lock:
        _pending[key] = {"event": ev, "reply": None}

    ev.wait(timeout=timeout)

    with _tg_lock:
        state = _pending.pop(key, {})

    reply = state.get("reply") or ""
    return jsonify({"reply": reply, "timed_out": not bool(reply)})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=9530, debug=False)
